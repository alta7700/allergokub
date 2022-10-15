from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from datetime import datetime
from main.models import Allergen, PollenCalendar


def parse_and_fill():
    wb = load_workbook(Path(__file__).parent / 'Календарь 2018-2022.04.06.xlsx')
    calendar = []
    for sheet in wb.worksheets:
        cell = sheet.cell(3, 1)
        allergens = {}
        while cell.value:
            allergens[cell.row] = Allergen.objects.get(title=cell.value.strip())
            cell = sheet.cell(cell.row + 1, 1)

        year = int(sheet.title)
        month = 3
        cell = sheet.cell(1, 2)
        while cell.value or (type(cell) == MergedCell):
            if cell.value:
                month = cell.value
            date = datetime(year=year, month=month, day=int(sheet.cell(2, cell.column).value)).date()
            for row, obj in allergens.items():
                try:
                    concentration = float(sheet.cell(row, cell.column).value or 0)
                except Exception as e:
                    print(date)
                    raise e
                calendar.append({'date': date, 'allergen': obj, 'concentration': concentration})
            cell = sheet.cell(1, cell.column + 1)
    calendar.sort(key=lambda x: x['date'])
    PollenCalendar.objects.bulk_create([PollenCalendar(**x) for x in calendar])
