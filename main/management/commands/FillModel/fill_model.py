from openpyxl import load_workbook


def fill(file, model, sort_field=None):
    wb = load_workbook(file)
    sheet = wb.active
    cell, fields = sheet.cell(1, 1), {}
    while cell.value:
        fields[cell.column] = cell.value
        cell = sheet.cell(1, cell.column + 1)
    cell, obj, objects_list = sheet.cell(2, 1), {}, []
    while cell.value:
        for num, field in fields.items():
            obj[field] = sheet.cell(cell.row, num).value
        cell = sheet.cell(cell.row + 1, 1)
        objects_list.append(obj)
        obj = {}
    if sort_field:
        objects_list.sort(key=lambda x: x[sort_field])
    model.objects.bulk_create(model(**x) for x in objects_list)
