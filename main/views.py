from datetime import datetime
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework.views import APIView
from rest_framework.generics import ListAPIView
from rest_framework.response import Response
from django.views.generic import FormView
from django.core.cache import cache
from django.http import HttpResponseRedirect, HttpResponse
from .models import TeamMember, Allergen, PollenCalendar
from .serializers import TeamMembersListSerializer, AllergensSerializer
from .forms import PollenAddDayForm
from .mixins import ResponseCacheMixin


class ImportCalendarDayView(FormView):
    template_name = 'main/import_calendar_day.html'
    success_url = '/admin/main/pollencalendar/'
    form_class = PollenAddDayForm

    def get_initial(self):
        return {'calendar_date': datetime.now().strftime('%Y-%m-%d')}

    def get(self, request, *args, **kwargs):
        if 'main.add_pollencalendar' in request.user.get_all_permissions():
            if 'get_excel_template' in request.GET:
                return HttpResponse(content=self.get_excel_template(), content_type='application/ms-excel',
                                    headers={'Content-Disposition': f'attachment; filename=Allergeni.xlsx'},
                                    )
            return self.render_to_response(self.get_context_data(**kwargs))
        return

    def form_valid(self, form):
        PollenCalendar.objects.bulk_create(
            [PollenCalendar(**x) for x in form.cleaned_data['values']]
        )
        cache.clear()
        return HttpResponseRedirect(self.get_success_url())

    @staticmethod
    def get_excel_template():
        book = Workbook()
        sheet = book.active
        sheet.cell(1, 1).value = 'Таксон'
        sheet.cell(1, 2).value = 'Концентрация'
        for i, allergen in enumerate(Allergen.objects.all()):
            sheet.cell(i+2, 1).value = allergen.title
            sheet.cell(i+2, 2).value = 0
        return save_virtual_workbook(book)


class TeamMembersAPIView(ResponseCacheMixin, ListAPIView):
    queryset = TeamMember.objects.filter(show=True)
    serializer_class = TeamMembersListSerializer

    cache_key_title = 'team_members'

    def get_data(self, request, *args, **kwargs):
        return self.get_serializer(self.filter_queryset(self.get_queryset()), many=True).data


class AllergenListAPIView(ResponseCacheMixin, ListAPIView):
    queryset = Allergen.objects.prefetch_related('calendar')
    serializer_class = AllergensSerializer

    cache_key_title = 'allergens_list'

    def get_data(self, request, *args, **kwargs):

        allergens = {}
        for obj in self.get_serializer(self.filter_queryset(self.get_queryset()), many=True).data:
            allergens[obj['id']] = obj['obj']
        return allergens


class AllergoMetrAPIView(ResponseCacheMixin, APIView):

    cache_key_title = 'current_common_risk'

    def get(self, request, *args, **kwargs):
        kwargs['get_data_only'] = True
        response_data = super().get(request, *args, **kwargs)
        if request.user.is_authenticated:
            response_data['self_risk'], response_data['user_allergens'] = self.get_self_risk(
                request.user, response_data['allergens']
            )
        else:
            response_data['self_risk'] = {}
            response_data['user_allergens'] = []
        return Response(response_data, status=200)

    def get_data(self, request, *args, **kwargs):
        last_values, actual_date = PollenCalendar.max_date_concentrations(('allergen',))
        response_body = {}
        max_score = 0
        allergens = {}
        # тип аллергена 1 (пыльца) или 2 (споры)
        common_concentration = {x[0]: [x[1], 0] for x in Allergen.TYPES}
        value: PollenCalendar
        for value in last_values:
            allergen = value.allergen
            score = allergen.get_risk_score(value.concentration)
            common_concentration[allergen.allergen_type][1] += value.concentration
            max_score += 4 * allergen.multiplier
            allergens[allergen.id] = {
                'allergen_type': allergen.allergen_type,
                'title': allergen.title,
                'score': score,
                'multiplier': allergen.multiplier,
                'abs_concentration': round(value.concentration, 2),
            }
        scores = sorted(list(allergens.values()), key=lambda x: x['score']*x['multiplier'], reverse=True)

        common_risk = {
            'current_risk': round(sum(x['score']*x['multiplier'] for x in scores) * 100 / max_score),
            'max_scores': ', '.join([x['title'] for x in scores[:2] if x['score'] != 0]),
            'pollen': {t_title: t_conc for t_title, t_conc in common_concentration.values()},
        }
        response_body['common_risk'] = common_risk
        response_body['actual_date'] = actual_date.strftime('%d.%m.%Y')
        response_body['allergens'] = allergens

        return response_body

    @classmethod
    def get_self_risk(cls, user, allergens):
        user_allergens = [str(x['id']) for x in user.allergens.values('id')]
        if not user_allergens:
            return {}, []
        scores = []
        max_score = 0
        for x in user_allergens:
            current_allergen = allergens[int(x)]
            scores.append(current_allergen)
            max_score += 4

        scores.sort(key=lambda x: x['score'], reverse=True)

        self_risk = {
            'current_risk': round(sum(x['score'] for x in scores) * 100 / max_score),
            'max_scores': ', '.join([x['title'] for x in scores[:2] if x['score'] != 0]),
        }
        return self_risk, user_allergens


class DateRangePollenAPIView(APIView):

    def get(self, request, *args, **kwargs):
        allergen = request.GET.get('allergen')
        start_date = request.GET.get('start')
        end_date = request.GET.get('end')
        response = {}
        if allergen and start_date and end_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.strptime(end_date, '%Y-%m-%d')
            response = PollenCalendar.date_range_concentration(allergen, start_date, end_date)
        return Response(response, status=200)
