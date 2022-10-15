from openpyxl import load_workbook
from django import forms
from .models import PollenCalendar, Allergen


class PollenAddDayForm(forms.Form):
    calendar_date = forms.DateField(label='Дата', widget=forms.DateInput(attrs={'type': 'date'}))
    file = forms.FileField(label='Файл excel')

    class Meta:
        fields = 'calendar_date', 'file'

    def clean(self):
        cleaned_data = super(PollenAddDayForm, self).clean()
        date = cleaned_data['calendar_date']
        wb = load_workbook(cleaned_data['file'])
        sheet = wb.active
        cell = sheet.cell(2, 1)
        values = {}
        allergens = {x.title: x for x in Allergen.objects.all()}
        while cell.value:
            try:
                allergen = allergens.pop(cell.value.strip())
            except KeyError:
                try:
                    Allergen.objects.get(title=cell.value.strip())
                    self.add_error('file', forms.ValidationError(f'Дубль аллергена {cell.value} (строка {cell.row})'))
                except Allergen.DoesNotExist:
                    self.add_error('file', forms.ValidationError(f'Нет аллергена {cell.value} (строка {cell.row})'))
                cell = sheet.cell(cell.row + 1, 1)
                continue
            except (AttributeError, Exception):
                self.add_error('file', forms.ValidationError(f'Некорректное название (строка {cell.row})'))
                cell = sheet.cell(cell.row + 1, 1)
                continue
            try:
                concentration = float(sheet.cell(cell.row, 2).value)
                if concentration < 0:
                    self.add_error(
                        'file',
                        forms.ValidationError(f'Отрицательная концентрация у "{allergen.title}" (строка {cell.row})')
                    )
                    continue
            except:
                self.add_error(
                    'file', forms.ValidationError(f'Некорректная концентрация у "{allergen.title}" (строка {cell.row})')
                )
                cell = sheet.cell(cell.row + 1, 1)
                continue
            values[allergen.id] = {'date': date, 'allergen': allergen, 'concentration': concentration}
            cell = sheet.cell(cell.row + 1, 1)

        if allergens:
            self.add_error('file', forms.ValidationError(f'Нет данных по: {", ".join(list(allergens.keys()))}'))

        this_date_report = list(PollenCalendar.objects.filter(date=date))
        if this_date_report:
            this_date_allergens = [x.allergen.id for x in this_date_report]
            date_repeat_allergens = []
            for allergen_id, note in values.items():
                if allergen_id in this_date_allergens:
                    date_repeat_allergens.append(note['allergen'].title)
            if date_repeat_allergens:
                date_repeat_allergens = ', '.join(date_repeat_allergens)
                self.add_error('calendar_date', forms.ValidationError(
                    f'За {date.strftime("%d.%m.%Y")} уже есть запись для: {date_repeat_allergens}'
                ))

        self.cleaned_data = {**cleaned_data, 'values': tuple(values.values())}
        return self.cleaned_data
